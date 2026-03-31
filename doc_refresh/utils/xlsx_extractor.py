"""
XLSX Sheet Extraction Utilities.

Provides functions for extracting individual sheets from Excel workbooks as
markdown text and as standalone xlsx files. Used by Stage 1 (scan) for sheet
discovery and by Stage 2 (extract) for content extraction and output file
generation. Designed to support the pipeline's treatment of each worksheet
as an independent document.

Key functions:
    get_xlsx_sheet_names: Discover non-empty sheets in a workbook.
    xlsx_sheet_to_markdown: Convert a single sheet to a markdown table.
    create_single_sheet_xlsx: Extract one sheet into a new workbook file.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from .token_utils import count_tokens, truncate_to_tokens

logger = logging.getLogger(__name__)

SHEET_NAME_UNSAFE = re.compile(r'[/\\:*?"<>|]')
MAX_SHEET_TOKENS = 80_000


def _normalize_cell(value: Any) -> str:
    """Normalize cell values to stable string content."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.15g}"
    return str(value).strip()


def _escape_markdown_pipe(text: str) -> str:
    """Escape pipe characters for use inside markdown table cells."""
    return text.replace("|", "\\|")


def sanitize_sheet_name(name: str) -> str:
    """Replace filesystem-unsafe characters in a sheet name with underscores.

    Args:
        name: Raw worksheet name from openpyxl.

    Returns:
        Sanitized string safe for use in file paths.
    """
    sanitized = SHEET_NAME_UNSAFE.sub("_", name).strip()
    return sanitized or "sheet"


def get_xlsx_sheet_names(xlsx_path: str) -> List[str]:
    """Return names of non-empty sheets in a workbook.

    A sheet is considered non-empty if it has at least one row with a
    non-None cell value. Empty sheets and sheets with only blank rows
    are excluded.

    Args:
        xlsx_path: Path to the xlsx file.

    Returns:
        List of sheet names that contain data.
    """
    workbook = load_workbook(
        filename=xlsx_path, data_only=True, read_only=True
    )
    try:
        names: List[str] = []
        for worksheet in workbook.worksheets:
            has_data = False
            for row in worksheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    has_data = True
                    break
            if has_data:
                names.append(worksheet.title)
        return names
    finally:
        workbook.close()


def _build_merged_cell_map(
    worksheet: Any,
) -> Dict[Tuple[int, int], Any]:
    """Build a map from merged cell coordinates to their top-left value."""
    merged_map: Dict[Tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        top_left_value = worksheet.cell(
            row=merged_range.min_row, column=merged_range.min_col
        ).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if row == merged_range.min_row and col == merged_range.min_col:
                    continue
                merged_map[(row, col)] = top_left_value
    return merged_map


def _get_cell_value(
    worksheet: Any,
    row: int,
    col: int,
    merged_map: Dict[Tuple[int, int], Any],
) -> Any:
    """Get a cell's value, resolving merged cells to their top-left value."""
    cell = worksheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return merged_map.get((row, col))
    return cell.value


def xlsx_sheet_to_markdown(
    xlsx_path: str,
    sheet_name: str,
    workbook_label: str,
) -> str:
    """Convert a single worksheet to a markdown string.

    Renders the sheet as a markdown heading followed by a markdown table.
    The first non-empty row is treated as the header row. Merged cells
    are resolved to their top-left value. Pipe characters in cell values
    are escaped for markdown table compatibility.

    Args:
        xlsx_path: Path to the xlsx file.
        sheet_name: Name of the worksheet to extract.
        workbook_label: Human-readable label for the workbook (typically stem).

    Returns:
        Markdown string with heading and table content.
    """
    workbook = load_workbook(
        filename=xlsx_path, data_only=True, read_only=False
    )
    try:
        worksheet = workbook[sheet_name]
        merged_map = _build_merged_cell_map(worksheet)

        min_row = worksheet.min_row or 1
        max_row = worksheet.max_row or 1
        min_col = worksheet.min_column or 1
        max_col = worksheet.max_column or 1

        rows_data: List[List[str]] = []
        for row_idx in range(min_row, max_row + 1):
            row_values: List[str] = []
            for col_idx in range(min_col, max_col + 1):
                raw = _get_cell_value(worksheet, row_idx, col_idx, merged_map)
                row_values.append(_normalize_cell(raw))
            rows_data.append(row_values)

        if not rows_data:
            return f"# {workbook_label} - {sheet_name}\n\n*Empty sheet*\n"

        first_nonempty_idx: Optional[int] = None
        for i, row in enumerate(rows_data):
            if any(cell.strip() for cell in row):
                first_nonempty_idx = i
                break

        if first_nonempty_idx is None:
            return f"# {workbook_label} - {sheet_name}\n\n*Empty sheet*\n"

        headers = rows_data[first_nonempty_idx]
        data_rows = rows_data[first_nonempty_idx + 1:]

        escaped_headers = [
            _escape_markdown_pipe(h) if h.strip() else f"Column {i + 1}"
            for i, h in enumerate(headers)
        ]

        lines: List[str] = []
        lines.append(f"# {workbook_label} - {sheet_name}")
        lines.append("")
        lines.append("| " + " | ".join(escaped_headers) + " |")
        lines.append("| " + " | ".join("---" for _ in escaped_headers) + " |")

        for row in data_rows:
            if not any(cell.strip() for cell in row):
                continue
            padded = row + [""] * (len(headers) - len(row))
            escaped = [_escape_markdown_pipe(cell) for cell in padded[: len(headers)]]
            lines.append("| " + " | ".join(escaped) + " |")

        lines.append("")
        markdown = "\n".join(lines)

        token_count = count_tokens(markdown)
        if token_count > MAX_SHEET_TOKENS:
            logger.warning(
                "Sheet '%s' in %s has %d tokens, truncating to %d",
                sheet_name,
                workbook_label,
                token_count,
                MAX_SHEET_TOKENS,
            )
            markdown = truncate_to_tokens(markdown, MAX_SHEET_TOKENS)

        return markdown
    finally:
        workbook.close()


def create_single_sheet_xlsx(
    xlsx_path: str,
    sheet_name: str,
    output_path: str,
) -> None:
    """Extract a single sheet from a workbook into a new xlsx file.

    Copies cell values (not formulas) from the source sheet into a fresh
    workbook and saves it to the output path. Column widths from the
    source are preserved where available.

    Args:
        xlsx_path: Path to the source xlsx file.
        sheet_name: Name of the worksheet to extract.
        output_path: Destination path for the new single-sheet xlsx file.
    """
    source_wb = load_workbook(filename=xlsx_path, data_only=True, read_only=False)
    try:
        source_ws = source_wb[sheet_name]
        merged_map = _build_merged_cell_map(source_ws)

        dest_wb = Workbook()
        try:
            dest_ws = dest_wb.active
            dest_ws.title = sheet_name

            for row_idx, source_row in enumerate(
                source_ws.iter_rows(
                    min_row=source_ws.min_row,
                    max_row=source_ws.max_row,
                    min_col=source_ws.min_column,
                    max_col=source_ws.max_column,
                ),
                start=1,
            ):
                for col_idx, source_cell in enumerate(source_row, start=1):
                    value = source_cell.value
                    if isinstance(source_cell, MergedCell):
                        actual_row = (source_ws.min_row or 1) + row_idx - 1
                        actual_col = (source_ws.min_column or 1) + col_idx - 1
                        value = merged_map.get((actual_row, actual_col))
                    dest_ws.cell(row=row_idx, column=col_idx, value=value)

            for col_letter, dim in source_ws.column_dimensions.items():
                if dim.width:
                    dest_ws.column_dimensions[col_letter].width = dim.width

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            dest_wb.save(output_path)
        finally:
            dest_wb.close()
    finally:
        source_wb.close()
