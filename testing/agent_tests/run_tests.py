#!/usr/bin/env python3
"""
IRIS Agent Test Runner

A lightweight testing framework for testing IRIS agents individually.
Reads test cases from Excel, executes agents directly, and outputs results to Excel.

Usage:
    python testing/agent_tests/run_tests.py
    python testing/agent_tests/run_tests.py --input custom_tests.xlsx
    python testing/agent_tests/run_tests.py --agents Router,Clarifier
    python testing/agent_tests/run_tests.py --tags regression
    python testing/agent_tests/run_tests.py --verbose
"""

import argparse
import json
import logging
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple

# =============================================================================
# PATH SETUP - Must happen before any iris imports
# =============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
sys.path.insert(0, PROJECT_ROOT)


# =============================================================================
# ENVIRONMENT SETUP - Must happen before iris imports
# =============================================================================


def setup_environment():
    """Configure environment for local OpenAI testing."""
    os.environ["AZURE_BASE_URL"] = "https://api.openai.com/v1"
    os.environ["IRIS_MODEL_SMALL"] = "gpt-4.1-mini"
    os.environ["IRIS_MODEL_LARGE"] = "gpt-4.1-mini"
    os.environ["IRIS_MODEL_EMBEDDING"] = "text-embedding-3-large"
    os.environ["IRIS_LOG_LEVEL"] = "WARNING"

    # Database config for local testing
    # Use current OS user as default (local postgres uses peer authentication)
    import getpass
    current_user = getpass.getuser()

    os.environ.setdefault("VECTOR_POSTGRES_DB_HOST", "localhost")
    os.environ.setdefault("VECTOR_POSTGRES_DB_PORT", "34532")
    os.environ.setdefault("VECTOR_POSTGRES_DB_NAME", "maven-finance")
    os.environ.setdefault("VECTOR_POSTGRES_DB_USERNAME", current_user)
    os.environ.setdefault("VECTOR_POSTGRES_DB_PASSWORD", "")


# Setup environment before imports
setup_environment()


# =============================================================================
# MONKEY-PATCH DATABASE CONNECTION FOR LOCAL TESTING (NO SSL)
# =============================================================================

from services.src.connections import postgres as db_config


def build_database_dsn_no_ssl(params: dict, for_sqlalchemy=True):
    """Modified DSN builder that disables SSL for local postgres."""
    host = params.get("host", "localhost")
    hosts = [h.strip() for h in str(host).split(",") if h.strip()]
    raw_port = str(params.get("port", "5432")).strip()
    ports = [p.strip() for p in raw_port.split(",") if p.strip()] if "," in raw_port else [raw_port] * len(hosts)
    dbname = params.get("dbname", "postgres")
    user = params.get("user", "postgres")
    password = params.get("password", "")

    if for_sqlalchemy:
        if len(hosts) == 1:
            return (
                f"postgresql+psycopg2://{user}:{password}@{hosts[0]}:{ports[0]}/{dbname}"
                f"?sslmode=disable&target_session_attrs=read-write"
            )
        hosts_str = ",".join(hosts)
        ports_str = ",".join(ports)
        return (
            f"postgresql+psycopg2://{user}:{password}@/{dbname}"
            f"?host={hosts_str}&port={ports_str}"
            f"&sslmode=disable&target_session_attrs=read-write"
        )
    else:
        hosts_str = ",".join(hosts)
        ports_str = ",".join(ports)
        return (
            f"host='{hosts_str}' port='{ports_str}' sslmode='disable' "
            f"target_session_attrs='read-write' dbname='{dbname}' user='{user}' password='{password}'"
        )


# Apply monkey-patch
db_config.build_database_dsn = build_database_dsn_no_ssl

# Now we can import iris modules
from services.src.utils.logging_format import configure_root_logger

configure_root_logger()

logger = logging.getLogger(__name__)

# Excel library
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# DATA STRUCTURES
# =============================================================================


class TestStatus(Enum):
    """Test result status."""

    PASS = "PASS"
    FAIL = "FAIL"
    ERROR = "ERROR"
    SKIP = "SKIP"


@dataclass
class TestCase:
    """Universal test case structure matching Excel columns."""

    test_id: str
    name: str
    tags: str
    conversation: Optional[str]  # JSON string
    research_statement: Optional[str]
    db_source: Optional[str]
    document_ids: Optional[str]  # Comma-separated
    aggregated_research: Optional[str]  # JSON string
    expected_decision: Optional[str]
    expected_action: Optional[str]
    expected_databases: Optional[str]  # Comma-separated
    expected_contains: Optional[str]  # Comma-separated
    expected_status: Optional[str]
    golden_response: Optional[str]
    min_score: Optional[float]
    notes: Optional[str]


@dataclass
class TestResult:
    """Result from running a single test."""

    test_id: str
    name: str
    status: TestStatus
    expected: str
    actual: str
    duration_ms: int
    error: Optional[str] = None
    validation_details: str = ""
    llm_score: Optional[float] = None
    actual_response_preview: str = ""
    usage_details: Dict[str, Any] = field(default_factory=dict)


@dataclass
class AgentResults:
    """Results for a single agent."""

    agent_name: str
    total: int = 0
    passed: int = 0
    failed: int = 0
    errors: int = 0
    skipped: int = 0
    total_duration_ms: int = 0
    test_results: List[TestResult] = field(default_factory=list)


# =============================================================================
# COLUMN DEFINITIONS - Agent-specific schemas
# =============================================================================

AGENT_COLUMNS = {
    "Router": ["test_id", "name", "conversation", "expected_decision", "notes"],
    "Clarifier": ["test_id", "name", "conversation", "expected_action", "notes"],
    "Planner": ["test_id", "name", "research_statement", "expected_databases", "notes"],
    "DirectResponse": ["test_id", "name", "conversation", "golden_response", "notes"],
    "Summarizer": ["test_id", "name", "research_statement", "aggregated_research", "golden_response", "notes"],
}

AGENT_SHEETS = list(AGENT_COLUMNS.keys())


# =============================================================================
# EXCEL READER
# =============================================================================


def load_test_cases(xlsx_path: str) -> Dict[str, List[TestCase]]:
    """
    Load test cases from Excel workbook.

    Args:
        xlsx_path: Path to the test cases Excel file.

    Returns:
        Dict mapping agent name to list of TestCase objects.
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Test cases file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    test_cases: Dict[str, List[TestCase]] = {}

    for sheet_name in AGENT_SHEETS:
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook", sheet_name)
            continue

        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            logger.warning("Sheet '%s' has no data rows", sheet_name)
            continue

        # First row is headers
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]

        # Map column indices based on agent-specific columns
        agent_columns = AGENT_COLUMNS.get(sheet_name, [])
        col_indices = {}
        for col_name in agent_columns:
            try:
                col_indices[col_name] = headers.index(col_name)
            except ValueError:
                col_indices[col_name] = -1

        cases = []
        for row_num, row in enumerate(rows[1:], start=2):
            # Skip empty rows
            if not row or all(cell is None for cell in row):
                continue

            def get_val(col_name: str) -> Any:
                idx = col_indices.get(col_name, -1)
                if idx < 0 or idx >= len(row):
                    return None
                return row[idx]

            test_id = get_val("test_id")
            if not test_id:
                continue

            case = TestCase(
                test_id=str(test_id),
                name=str(get_val("name") or ""),
                tags=str(get_val("tags") or ""),
                conversation=str(get_val("conversation")) if get_val("conversation") else None,
                research_statement=str(get_val("research_statement")) if get_val("research_statement") else None,
                db_source=str(get_val("db_source")) if get_val("db_source") else None,
                document_ids=str(get_val("document_ids")) if get_val("document_ids") else None,
                aggregated_research=str(get_val("aggregated_research")) if get_val("aggregated_research") else None,
                expected_decision=str(get_val("expected_decision")) if get_val("expected_decision") else None,
                expected_action=str(get_val("expected_action")) if get_val("expected_action") else None,
                expected_databases=str(get_val("expected_databases")) if get_val("expected_databases") else None,
                expected_contains=str(get_val("expected_contains")) if get_val("expected_contains") else None,
                expected_status=str(get_val("expected_status")) if get_val("expected_status") else None,
                golden_response=str(get_val("golden_response")) if get_val("golden_response") else None,
                min_score=float(get_val("min_score")) if get_val("min_score") else None,
                notes=str(get_val("notes")) if get_val("notes") else None,
            )
            cases.append(case)

        if cases:
            test_cases[sheet_name] = cases
            logger.info("Loaded %d test cases from sheet '%s'", len(cases), sheet_name)

    wb.close()
    return test_cases


def filter_by_tags(cases: List[TestCase], tags: List[str]) -> List[TestCase]:
    """Filter test cases by tags (case must have ALL specified tags)."""
    if not tags:
        return cases

    filtered = []
    for case in cases:
        case_tags = [t.strip().lower() for t in case.tags.split(",") if t.strip()]
        if all(t.lower() in case_tags for t in tags):
            filtered.append(case)
    return filtered


# =============================================================================
# AGENT EXECUTION FUNCTIONS
# =============================================================================


def fetch_available_databases() -> Dict[str, Any]:
    """Get available databases from registry."""
    try:
        from services.src.agent.tools.database_metadata import (
            fetch_available_databases as fetch_dbs,
        )

        return fetch_dbs()
    except Exception as e:
        logger.warning("Could not load databases: %s", e)
        return {}


def parse_conversation(json_str: str) -> Dict[str, Any]:
    """Parse conversation JSON string."""
    try:
        messages = json.loads(json_str)
        return {"messages": messages}
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid conversation JSON: {e}")


def run_router_test(case: TestCase, api_key: str) -> TestResult:
    """Execute Router agent test."""
    from services.src.agent.router import generate_routing_decision

    start_time = time.time()
    error = None
    actual = ""
    validation_details = ""

    try:
        if not case.conversation:
            raise ValueError("Router test requires 'conversation' field")

        conversation = parse_conversation(case.conversation)
        available_databases = fetch_available_databases()

        decision, usage = generate_routing_decision(
            conversation=conversation,
            token=api_key,
            available_databases=available_databases,
        )

        actual = decision.get("function_name", "")

        # Validate
        expected = case.expected_decision or ""
        if actual.lower() == expected.lower():
            status = TestStatus.PASS
            validation_details = f"exact_match: {actual} == {expected}"
        else:
            status = TestStatus.FAIL
            validation_details = f"expected '{expected}' but got '{actual}'"

    except Exception as e:
        status = TestStatus.ERROR
        error = str(e)
        logger.error("Router test error: %s", e, exc_info=True)

    duration_ms = int((time.time() - start_time) * 1000)

    return TestResult(
        test_id=case.test_id,
        name=case.name,
        status=status,
        expected=case.expected_decision or "",
        actual=actual,
        duration_ms=duration_ms,
        error=error,
        validation_details=validation_details,
    )


def run_clarifier_test(case: TestCase, api_key: str) -> TestResult:
    """Execute Clarifier agent test."""
    from services.src.agent.clarifier import generate_clarifier_decision

    start_time = time.time()
    error = None
    actual = ""
    validation_details = ""

    try:
        if not case.conversation:
            raise ValueError("Clarifier test requires 'conversation' field")

        conversation = parse_conversation(case.conversation)
        available_databases = fetch_available_databases()

        decision, usage = generate_clarifier_decision(
            conversation=conversation,
            token=api_key,
            available_databases=available_databases,
        )

        actual_action = decision.get("action", "")
        actual = actual_action

        # Validate action
        expected_action = case.expected_action or ""
        action_match = actual_action.lower() == expected_action.lower()

        if action_match:
            status = TestStatus.PASS
            validation_details = f"exact_match: {actual_action} == {expected_action}"
        else:
            status = TestStatus.FAIL
            validation_details = f"expected '{expected_action}' but got '{actual_action}'"

    except Exception as e:
        status = TestStatus.ERROR
        error = str(e)
        logger.error("Clarifier test error: %s", e, exc_info=True)

    duration_ms = int((time.time() - start_time) * 1000)

    return TestResult(
        test_id=case.test_id,
        name=case.name,
        status=status,
        expected=case.expected_action or "",
        actual=actual,
        duration_ms=duration_ms,
        error=error,
        validation_details=validation_details,
    )


def run_planner_test(case: TestCase, api_key: str) -> TestResult:
    """Execute Planner agent test."""
    from services.src.agent.planner import generate_database_selection_plan

    start_time = time.time()
    error = None
    actual = ""
    validation_details = ""

    try:
        if not case.research_statement:
            raise ValueError("Planner test requires 'research_statement' field")

        available_databases = fetch_available_databases()

        plan, usage_list = generate_database_selection_plan(
            research_statement=case.research_statement,
            token=api_key,
            available_databases=available_databases,
        )

        actual_dbs = plan.get("databases", [])
        actual = ",".join(actual_dbs)

        # Validate expected databases
        expected_dbs = []
        if case.expected_databases:
            expected_dbs = [d.strip().lower() for d in case.expected_databases.split(",") if d.strip()]

        actual_dbs_lower = [d.lower() for d in actual_dbs]
        all_expected_found = all(exp in actual_dbs_lower for exp in expected_dbs)

        if all_expected_found:
            status = TestStatus.PASS
            validation_details = f"expected DBs found: {expected_dbs} in {actual_dbs}"
        else:
            status = TestStatus.FAIL
            validation_details = f"expected '{expected_dbs}' but got '{actual_dbs}'"

    except Exception as e:
        status = TestStatus.ERROR
        error = str(e)
        logger.error("Planner test error: %s", e, exc_info=True)

    duration_ms = int((time.time() - start_time) * 1000)

    return TestResult(
        test_id=case.test_id,
        name=case.name,
        status=status,
        expected=case.expected_databases or "",
        actual=actual,
        duration_ms=duration_ms,
        error=error,
        validation_details=validation_details,
    )


def run_direct_response_test(case: TestCase, api_key: str) -> TestResult:
    """Execute DirectResponse agent test."""
    from services.src.agent.direct_response import (
        stream_direct_response_from_conversation,
    )

    start_time = time.time()
    error = None
    actual = ""
    validation_details = ""
    llm_score = None
    response_text = ""

    try:
        if not case.conversation:
            raise ValueError("DirectResponse test requires 'conversation' field")

        conversation = parse_conversation(case.conversation)
        available_databases = fetch_available_databases()

        # Collect streaming response
        usage_details = {}
        for chunk in stream_direct_response_from_conversation(
            conversation=conversation,
            token=api_key,
            available_databases=available_databases,
        ):
            if isinstance(chunk, dict) and "usage_details" in chunk:
                usage_details = chunk["usage_details"]
            elif isinstance(chunk, str):
                response_text += chunk

        actual = response_text[:100] + "..." if len(response_text) > 100 else response_text

        # LLM Judge if golden_response provided (default min_score: 0.7)
        if case.golden_response:
            min_score = case.min_score if case.min_score else 0.7
            llm_score = run_llm_judge(response_text, case.golden_response, api_key)
            score_pass = llm_score >= min_score
            status = TestStatus.PASS if score_pass else TestStatus.FAIL
            validation_details = f"llm_score: {llm_score:.2f} >= {min_score}"
        else:
            # No golden_response - just check response was generated
            status = TestStatus.PASS if response_text else TestStatus.FAIL
            validation_details = f"response_generated: {bool(response_text)}"

    except Exception as e:
        status = TestStatus.ERROR
        error = str(e)
        logger.error("DirectResponse test error: %s", e, exc_info=True)

    duration_ms = int((time.time() - start_time) * 1000)

    return TestResult(
        test_id=case.test_id,
        name=case.name,
        status=status,
        expected=case.golden_response[:50] + "..." if case.golden_response else "response generated",
        actual=actual,
        duration_ms=duration_ms,
        error=error,
        validation_details=validation_details,
        llm_score=llm_score,
        actual_response_preview=response_text[:500],
    )


def run_summarizer_test(case: TestCase, api_key: str) -> TestResult:
    """Execute Summarizer agent test."""
    from services.src.agent.summarizer import stream_research_summary

    start_time = time.time()
    error = None
    actual = ""
    validation_details = ""
    llm_score = None
    response_text = ""

    try:
        if not case.aggregated_research:
            raise ValueError("Summarizer test requires 'aggregated_research' field")

        aggregated_research = json.loads(case.aggregated_research)
        available_databases = fetch_available_databases()

        summary_context = None
        if case.research_statement:
            summary_context = {"research_statement": case.research_statement}

        # Collect streaming response
        for chunk in stream_research_summary(
            aggregated_detailed_research=aggregated_research,
            token=api_key,
            available_databases=available_databases,
            summary_context=summary_context,
        ):
            if isinstance(chunk, dict) and "usage_details" in chunk:
                pass  # Usage details
            elif isinstance(chunk, str):
                response_text += chunk

        actual = response_text[:100] + "..." if len(response_text) > 100 else response_text

        # LLM Judge if golden_response provided (default min_score: 0.7)
        if case.golden_response:
            min_score = case.min_score if case.min_score else 0.7
            llm_score = run_llm_judge(response_text, case.golden_response, api_key)
            score_pass = llm_score >= min_score
            status = TestStatus.PASS if score_pass else TestStatus.FAIL
            validation_details = f"llm_score: {llm_score:.2f} >= {min_score}"
        else:
            # No golden_response - just check response was generated
            status = TestStatus.PASS if response_text else TestStatus.FAIL
            validation_details = f"response_generated: {bool(response_text)}"

    except Exception as e:
        status = TestStatus.ERROR
        error = str(e)
        logger.error("Summarizer test error: %s", e, exc_info=True)

    duration_ms = int((time.time() - start_time) * 1000)

    return TestResult(
        test_id=case.test_id,
        name=case.name,
        status=status,
        expected=case.golden_response[:50] + "..." if case.golden_response else "response generated",
        actual=actual,
        duration_ms=duration_ms,
        error=error,
        validation_details=validation_details,
        llm_score=llm_score,
        actual_response_preview=response_text[:500],
    )


def run_metadata_subagent_test(case: TestCase, api_key: str) -> TestResult:
    """Execute MetadataSubagent test."""
    from services.src.agent.tools.metadata_subagent import (
        execute_unified_metadata_query,
    )

    start_time = time.time()
    error = None
    actual = ""
    validation_details = ""

    try:
        if not case.research_statement or not case.db_source:
            raise ValueError("MetadataSubagent test requires 'research_statement' and 'db_source' fields")

        # Build query context
        query_context = {
            "token": api_key,
            "process_monitor": None,
            "stage_name": "test",
            "query_embedding": None,  # Will be generated
        }

        result = execute_unified_metadata_query(
            research_statement=case.research_statement,
            db_source=case.db_source,
            query_context=query_context,
        )

        # findings contains all relevant results (answered + needs_deep_research)
        findings_count = len(result.get("findings", []))
        needs_research_count = len(result.get("needs_research_doc_ids", []))
        irrelevant_count = result.get("irrelevant_count", 0)

        actual = f"findings:{findings_count}, needs_research:{needs_research_count}, irrelevant:{irrelevant_count}"

        # Validate expected status
        expected_status = (case.expected_status or "success").lower()
        has_results = findings_count > 0 or needs_research_count > 0

        if expected_status == "success" and has_results:
            status = TestStatus.PASS
            validation_details = f"Found results: {actual}"
        elif expected_status == "error" and not has_results:
            status = TestStatus.PASS
            validation_details = "Expected no results and got none"
        else:
            status = TestStatus.FAIL
            validation_details = f"Expected '{expected_status}', got results: {has_results}"

    except Exception as e:
        status = TestStatus.ERROR
        error = str(e)
        logger.error("MetadataSubagent test error: %s", e, exc_info=True)

    duration_ms = int((time.time() - start_time) * 1000)

    return TestResult(
        test_id=case.test_id,
        name=case.name,
        status=status,
        expected=case.expected_status or "success",
        actual=actual,
        duration_ms=duration_ms,
        error=error,
        validation_details=validation_details,
    )


def run_file_research_test(case: TestCase, api_key: str) -> TestResult:
    """Execute FileResearch subagent test."""
    from services.src.agent.tools.file_research_subagent import execute_file_research_sync

    start_time = time.time()
    error = None
    actual = ""
    validation_details = ""

    try:
        if not case.research_statement or not case.db_source or not case.document_ids:
            raise ValueError("FileResearch test requires 'research_statement', 'db_source', and 'document_ids' fields")

        document_ids = [d.strip() for d in case.document_ids.split(",") if d.strip()]

        research_context = {
            "token": api_key,
            "process_monitor": None,
            "stage_name": "test",
        }

        result = execute_file_research_sync(
            research_statement=case.research_statement,
            document_ids=document_ids,
            db_source=case.db_source,
            research_context=research_context,
        )

        docs_processed = len(result.get("documents", {}))
        status_summary = result.get("status_summary", "")
        actual = f"docs_processed: {docs_processed}"

        # Validate expected status
        expected_status = (case.expected_status or "success").lower()

        if expected_status == "success" and docs_processed > 0:
            status = TestStatus.PASS
            validation_details = f"Processed {docs_processed} documents"
        elif expected_status == "error" and docs_processed == 0:
            status = TestStatus.PASS
            validation_details = "Expected error and got no results"
        else:
            status = TestStatus.FAIL
            validation_details = f"Expected '{expected_status}', processed {docs_processed} docs"

    except Exception as e:
        status = TestStatus.ERROR
        error = str(e)
        logger.error("FileResearch test error: %s", e, exc_info=True)

    duration_ms = int((time.time() - start_time) * 1000)

    return TestResult(
        test_id=case.test_id,
        name=case.name,
        status=status,
        expected=case.expected_status or "success",
        actual=actual,
        duration_ms=duration_ms,
        error=error,
        validation_details=validation_details,
    )


# =============================================================================
# LLM-AS-JUDGE VALIDATOR
# =============================================================================

LLM_JUDGE_SYSTEM = """You are an expert evaluator comparing AI responses.

Score the ACTUAL RESPONSE against the GOLDEN RESPONSE on these criteria:
- Relevance (0-1): Does it address the same topic?
- Accuracy (0-1): Is the information correct?
- Completeness (0-1): Does it cover key points?
- Coherence (0-1): Is it well-structured?

Return ONLY a JSON object:
{"relevance": 0.0, "accuracy": 0.0, "completeness": 0.0, "coherence": 0.0, "weighted_score": 0.0, "reasoning": "..."}

The weighted_score should be: (relevance*0.25 + accuracy*0.30 + completeness*0.25 + coherence*0.20)"""


def run_llm_judge(actual_response: str, golden_response: str, api_key: str) -> float:
    """
    Use LLM to judge response quality against golden response.

    Returns weighted score between 0.0 and 1.0.
    """
    from services.src.connections.llm import execute_llm_call

    try:
        messages = [
            {"role": "system", "content": LLM_JUDGE_SYSTEM},
            {
                "role": "user",
                "content": f"GOLDEN RESPONSE:\n{golden_response}\n\nACTUAL RESPONSE:\n{actual_response}",
            },
        ]

        response, _ = execute_llm_call(
            oauth_token=api_key,
            model="gpt-4.1-mini",
            messages=messages,
            max_tokens=500,
            temperature=0.0,
            stream=False,
            prompt_token_cost=0,
            completion_token_cost=0,
        )

        content = response.choices[0].message.content
        evaluation = json.loads(content)
        return float(evaluation.get("weighted_score", 0.0))

    except Exception as e:
        logger.error("LLM judge error: %s", e)
        return 0.0


# =============================================================================
# TEST RUNNER
# =============================================================================

AGENT_RUNNERS = {
    "Router": run_router_test,
    "Clarifier": run_clarifier_test,
    "Planner": run_planner_test,
    "DirectResponse": run_direct_response_test,
    "Summarizer": run_summarizer_test,
}


def run_tests(
    test_cases: Dict[str, List[TestCase]],
    api_key: str,
    agents_filter: Optional[List[str]] = None,
    tags_filter: Optional[List[str]] = None,
    verbose: bool = False,
) -> Dict[str, AgentResults]:
    """
    Run all tests and return results.

    Args:
        test_cases: Dict mapping agent name to test cases.
        api_key: OpenAI API key.
        agents_filter: Optional list of agent names to run.
        tags_filter: Optional list of tags to filter by.
        verbose: Print progress during execution.

    Returns:
        Dict mapping agent name to AgentResults.
    """
    results: Dict[str, AgentResults] = {}

    for agent_name, cases in test_cases.items():
        # Filter by agent
        if agents_filter and agent_name not in agents_filter:
            continue

        # Filter by tags
        if tags_filter:
            cases = filter_by_tags(cases, tags_filter)
            if not cases:
                continue

        runner = AGENT_RUNNERS.get(agent_name)
        if not runner:
            logger.warning("No runner for agent: %s", agent_name)
            continue

        agent_results = AgentResults(agent_name=agent_name)

        if verbose:
            print(f"\n{'='*60}")
            print(f"Running {agent_name} tests ({len(cases)} cases)")
            print("=" * 60)

        for case in cases:
            if verbose:
                print(f"  Running: {case.test_id} - {case.name}...", end=" ", flush=True)

            result = runner(case, api_key)
            agent_results.test_results.append(result)
            agent_results.total += 1
            agent_results.total_duration_ms += result.duration_ms

            if result.status == TestStatus.PASS:
                agent_results.passed += 1
                if verbose:
                    print(f"PASS ({result.duration_ms / 1000:.2f}s)")
            elif result.status == TestStatus.FAIL:
                agent_results.failed += 1
                if verbose:
                    print(f"FAIL ({result.duration_ms / 1000:.2f}s)")
            elif result.status == TestStatus.ERROR:
                agent_results.errors += 1
                if verbose:
                    print(f"ERROR ({result.duration_ms / 1000:.2f}s): {result.error}")
            else:
                agent_results.skipped += 1
                if verbose:
                    print("SKIP")

        results[agent_name] = agent_results

    return results


# =============================================================================
# EXCEL WRITER
# =============================================================================


def write_results_xlsx(results: Dict[str, AgentResults], output_path: str):
    """
    Write test results to Excel workbook.

    Creates:
    - Summary sheet with overall stats
    - Per-agent result sheets
    """
    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    error_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = ["Agent", "Total", "Passed", "Failed", "Errors", "Pass Rate", "Avg Duration (ms)"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row = 2
    total_all = passed_all = failed_all = errors_all = 0

    for agent_name, agent_results in results.items():
        total_all += agent_results.total
        passed_all += agent_results.passed
        failed_all += agent_results.failed
        errors_all += agent_results.errors

        pass_rate = f"{(agent_results.passed / agent_results.total * 100):.0f}%" if agent_results.total > 0 else "N/A"
        avg_duration = agent_results.total_duration_ms // agent_results.total if agent_results.total > 0 else 0

        values = [
            agent_name,
            agent_results.total,
            agent_results.passed,
            agent_results.failed,
            agent_results.errors,
            pass_rate,
            avg_duration,
        ]

        for col, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            cell.border = thin_border

        row += 1

    # Totals row
    pass_rate_all = f"{(passed_all / total_all * 100):.0f}%" if total_all > 0 else "N/A"
    totals = ["TOTAL", total_all, passed_all, failed_all, errors_all, pass_rate_all, "-"]
    for col, value in enumerate(totals, 1):
        cell = ws_summary.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.border = thin_border

    # Auto-width columns
    for col in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 15

    # Per-agent result sheets
    for agent_name, agent_results in results.items():
        ws = wb.create_sheet(title=f"{agent_name}_Results")

        headers = ["test_id", "name", "status", "expected", "actual", "duration_ms", "error", "validation_details"]
        if agent_name in ["DirectResponse", "Summarizer"]:
            headers.extend(["llm_score", "actual_response_preview"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_num, result in enumerate(agent_results.test_results, 2):
            values = [
                result.test_id,
                result.name,
                result.status.value,
                result.expected,
                result.actual,
                result.duration_ms,
                result.error or "",
                result.validation_details,
            ]

            if agent_name in ["DirectResponse", "Summarizer"]:
                values.extend([
                    result.llm_score if result.llm_score is not None else "",
                    result.actual_response_preview[:200] if result.actual_response_preview else "",
                ])

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border

                # Color status column
                if col == 3:
                    if result.status == TestStatus.PASS:
                        cell.fill = pass_fill
                    elif result.status == TestStatus.FAIL:
                        cell.fill = fail_fill
                    elif result.status == TestStatus.ERROR:
                        cell.fill = error_fill

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(output_path)
    logger.info("Results saved to: %s", output_path)


# =============================================================================
# MAIN CLI
# =============================================================================


def main():
    parser = argparse.ArgumentParser(description="IRIS Agent Test Runner")
    parser.add_argument(
        "--input",
        type=str,
        default=os.path.join(SCRIPT_DIR, "input", "test_cases.xlsx"),
        help="Path to test cases Excel file",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Path for output Excel file (default: output/results_TIMESTAMP.xlsx)",
    )
    parser.add_argument(
        "--agents",
        type=str,
        default=None,
        help="Comma-separated list of agents to test (e.g., Router,Clarifier)",
    )
    parser.add_argument(
        "--tags",
        type=str,
        default=None,
        help="Comma-separated list of tags to filter by",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress during execution",
    )

    args = parser.parse_args()

    # Check API key
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("ERROR: OPENAI_API_KEY environment variable not set")
        print("  export OPENAI_API_KEY='sk-...'")
        sys.exit(1)

    # Parse filters
    agents_filter = [a.strip() for a in args.agents.split(",")] if args.agents else None
    tags_filter = [t.strip() for t in args.tags.split(",")] if args.tags else None

    # Load test cases
    print(f"Loading test cases from: {args.input}")
    try:
        test_cases = load_test_cases(args.input)
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        print("Create the test cases Excel file or specify a different path with --input")
        sys.exit(1)

    if not test_cases:
        print("ERROR: No test cases found in workbook")
        sys.exit(1)

    total_cases = sum(len(cases) for cases in test_cases.values())
    print(f"Loaded {total_cases} test cases across {len(test_cases)} agents")

    # Run tests
    print(f"\nRunning tests...")
    results = run_tests(
        test_cases=test_cases,
        api_key=api_key,
        agents_filter=agents_filter,
        tags_filter=tags_filter,
        verbose=args.verbose,
    )

    # Generate output path
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(SCRIPT_DIR, "output", f"results_{timestamp}.xlsx")

    # Write results
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    write_results_xlsx(results, output_path)

    # Print summary
    print(f"\n{'='*60}")
    print("TEST SUMMARY")
    print("=" * 60)

    total_tests = total_passed = total_failed = total_errors = 0

    for agent_name, agent_results in results.items():
        total_tests += agent_results.total
        total_passed += agent_results.passed
        total_failed += agent_results.failed
        total_errors += agent_results.errors

        pass_rate = f"{(agent_results.passed / agent_results.total * 100):.0f}%" if agent_results.total > 0 else "N/A"
        print(f"  {agent_name}: {agent_results.passed}/{agent_results.total} passed ({pass_rate})")

    print("-" * 60)
    overall_rate = f"{(total_passed / total_tests * 100):.0f}%" if total_tests > 0 else "N/A"
    print(f"  TOTAL: {total_passed}/{total_tests} passed ({overall_rate})")
    print(f"\nResults saved to: {output_path}")

    # Exit code
    if total_failed > 0 or total_errors > 0:
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
